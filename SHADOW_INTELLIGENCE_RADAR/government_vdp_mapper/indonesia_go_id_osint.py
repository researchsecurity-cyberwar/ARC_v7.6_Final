import requests
from bs4 import BeautifulSoup
import re
import tempfile
import os

class IndonesiaGoIdOSINT:
    """
    .go.id + OJK fintech register + GitHub dorks.
    Mengumpulkan target VDP Indonesia melalui OSINT dari sumber yang terverifikasi aktif.
    """
    
    def __init__(self):
        # Sumber yang TERBUKTI AKTIF berdasarkan verifikasi lapangan
        self.go_id_sources = [
            "https://www.lkpp.go.id",           # ✅ Aktif - Daftar Pelaku Usaha & Instansi
            "https://data.go.id"                 # ✅ Aktif - Portal Satu Data Indonesia
        ]
        # URL OJK akan dicari secara dinamis karena sering berubah
        self.ojk_base_url = "https://ojk.go.id"
    
    def collect_go_id_domains(self):
        """Kumpulkan domain .go.id dari sumber publik yang aktif."""
        domains = set()
        
        for source_url in self.go_id_sources:
            try:
                print(f"🔍 Scraping {source_url}...")
                response = requests.get(source_url, timeout=10)
                if response.status_code == 200:
                    if "lkpp.go.id" in source_url:
                        found_domains = self._extract_from_lkpp(response.text)
                    elif "data.go.id" in source_url:
                        found_domains = self._extract_from_data_go_id(response.text)
                    else:
                        found_domains = self._generic_go_id_extraction(response.text)
                    
                    domains.update(found_domains)
                    print(f"✅ Found {len(found_domains)} domains from {source_url}")
            except Exception as e:
                print(f"⚠️ Failed to scrape {source_url}: {e}")
                continue
        
        return list(domains)
    
    def _extract_from_lkpp(self, html_content):
        """Ekstrak domain dari lkpp.go.id berdasarkan struktur aktual."""
        domains = set()
        
        # Dari struktur yang kamu berikan, lkpp.go.id punya kategori:
        # Pelaku Usaha, K/L, BUMN, dll
        # Kita cari pola umum .go.id di seluruh konten
        go_id_pattern = r'https?://[a-zA-Z0-9.-]*\.go\.id'
        found_domains = re.findall(go_id_pattern, html_content)
        
        for domain in found_domains:
            # Bersihkan dan validasi
            clean_domain = domain.split('://')[-1].split('/')[0].lower()
            if clean_domain.endswith('.go.id') and '*' not in clean_domain:
                domains.add(clean_domain)
        
        return list(domains)
    
    def _extract_from_data_go_id(self, html_content):
        """Ekstrak domain dari data.go.id berdasarkan struktur aktual."""
        domains = set()
        
        # Portal data.go.id menyebutkan "Kementerian/Lembaga", "Provinsi", dll
        # Cari pola .go.id dalam konten
        go_id_pattern = r'[a-zA-Z0-9.-]*\.go\.id'
        found_domains = re.findall(go_id_pattern, html_content)
        
        for domain in found_domains:
            if domain.endswith('.go.id') and '*' not in domain:
                domains.add(domain.lower())
        
        return list(domains)
    
    def _generic_go_id_extraction(self, html_content):
        """Ekstraksi generik untuk sumber lain."""
        domains = set()
        go_id_pattern = r'[a-zA-Z0-9.-]*\.go\.id'
        found_domains = re.findall(go_id_pattern, html_content)
        
        for domain in found_domains:
            if domain.endswith('.go.id') and '*' not in domain:
                domains.add(domain.lower())
        
        return list(domains)
    
    def find_current_ojk_fintech_page(self):
        """Cari halaman daftar fintech OJK yang aktif saat ini."""
        try:
            # Cari di halaman utama OJK
            response = requests.get(self.ojk_base_url, timeout=10)
            if response.status_code == 200:
                soup = BeautifulSoup(response.content, 'html.parser')
                
                # Cari link yang mengandung kata kunci fintech
                keywords = ['fintech', 'tekfin', 'financial technology', 'daftar']
                for keyword in keywords:
                    links = soup.find_all('a', href=True, 
                        string=re.compile(keyword, re.IGNORECASE))
                    for link in links:
                        href = link['href']
                        if href.startswith('/'):
                            full_url = f"{self.ojk_base_url}{href}"
                        elif href.startswith('http'):
                            full_url = href
                        else:
                            continue
                        
                        # Verifikasi apakah halaman tersebut aktif
                        if self._is_ojk_fintech_page_valid(full_url):
                            return full_url
        except Exception as e:
            print(f"⚠️ Failed to find OJK fintech page: {e}")
        
        # Fallback: Return None jika tidak ditemukan
        return None
    
    def _is_ojk_fintech_page_valid(self, url):
        """Verifikasi apakah halaman OJK berisi daftar fintech."""
        try:
            response = requests.get(url, timeout=5)
            if response.status_code == 200:
                content_lower = response.text.lower()
                # Cek keberadaan indikator daftar fintech
                indicators = ['terdaftar', 'berizin', 'fintech', 'penyelenggara']
                return any(indicator in content_lower for indicator in indicators)
        except:
            pass
        return False
    
    def get_ojk_fintech_list(self):
        """Dapatkan daftar fintech terdaftar dari OJK (dengan pencarian dinamis)."""
        ojk_fintech_url = self.find_current_ojk_fintech_page()
        
        if not ojk_fintech_url:
            print("⚠️ Could not find current OJK fintech page. Using fallback method.")
            return self._fallback_ojk_fintech_scraping()
        
        try:
            response = requests.get(ojk_fintech_url, timeout=10)
            if response.status_code == 200:
                soup = BeautifulSoup(response.content, 'html.parser')
                return self._extract_fintech_companies(soup)
        except Exception as e:
            print(f"⚠️ Failed to fetch OJK fintech list from {ojk_fintech_url}: {e}")
        
        return []
    
    def _extract_fintech_companies(self, soup):
        """Ekstrak daftar perusahaan fintech dari halaman OJK."""
        companies = []
        # Cari tabel atau daftar perusahaan
        table_rows = soup.find_all('tr')
        
        for row in table_rows[1:]:  # Skip header
            cells = row.find_all('td')
            if len(cells) >= 2:
                company_name = cells[0].get_text(strip=True)
                website = cells[1].get_text(strip=True)
                
                if website and company_name:
                    companies.append({
                        'name': company_name,
                        'website': website,
                        'category': 'fintech',
                        'regulator': 'OJK'
                    })
        
        return companies
    
    def _fallback_ojk_fintech_scraping(self):
        """Metode fallback untuk mendapatkan daftar fintech OJK."""
        # Gunakan Google dork untuk mencari daftar fintech OJK terbaru
        # Ini adalah strategi OSINT yang valid
        print("🔍 Using fallback OSINT strategy for OJK fintech...")
        
        # Kita kembalikan daftar kosong untuk sekarang
        # Nanti ARC bisa menggunakan GitHub dorks atau sumber lain
        return []
    
    def github_go_id_dorks(self):
        """Generate GitHub dorks untuk mencari secrets di repositori .go.id."""
        dorks = [
            'filename:.env site:github.com "*.go.id"',
            'filename:config site:github.com "*.go.id"',
            'password site:github.com "*.go.id"',
            'api_key site:github.com "*.go.id"',
            'secret site:github.com "*.go.id"',
            'filetype:yaml site:github.com "*.go.id"',
            'filetype:json site:github.com "*.go.id"'
        ]
        return dorks
    
    # =============== INTEGRASI PARSING DOKUMEN ===============
    def extract_pdf_content(self, pdf_url):
        """Ekstrak konten dari PDF dengan penanganan error yang aman."""
        try:
            import pdfplumber
            
            # Download ke file sementara
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_file:
                response = requests.get(pdf_url, timeout=15)
                tmp_file.write(response.content)
                tmp_path = tmp_file.name
            
            try:
                # Ekstrak teks dari PDF
                text = ""
                with pdfplumber.open(tmp_path) as pdf:
                    for page in pdf.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n"
                return text.strip()
            finally:
                # Hapus file sementara
                os.unlink(tmp_path)
                
        except ImportError:
            print("⚠️ pdfplumber not available. Install it with: pip install pdfplumber")
            return None
        except Exception as e:
            print(f"⚠️ PDF extraction failed for {pdf_url}: {e}")
            return None
    
    def extract_docx_content(self, docx_url):
        """Ekstrak konten dari DOCX dengan penanganan error yang aman."""
        try:
            from docx import Document
            
            # Download ke file sementara
            with tempfile.NamedTemporaryFile(suffix='.docx', delete=False) as tmp_file:
                response = requests.get(docx_url, timeout=15)
                tmp_file.write(response.content)
                tmp_path = tmp_file.name
            
            try:
                # Ekstrak teks dari DOCX
                doc = Document(tmp_path)
                text = "\n".join([para.text for para in doc.paragraphs])
                return text.strip()
            finally:
                os.unlink(tmp_path)
                
        except ImportError:
            print("⚠️ python-docx not available. Install it with: pip install python-docx")
            return None
        except Exception as e:
            print(f"⚠️ DOCX extraction failed for {docx_url}: {e}")
            return None
    
    def _download_and_extract_document(self, url):
        """Download dan ekstrak konten dari dokumen (PDF/DOCX/XLSX)."""
        if url.lower().endswith('.pdf'):
            return self.extract_pdf_content(url)
        elif url.lower().endswith('.docx'):
            return self.extract_docx_content(url)
        elif url.lower().endswith(('.xlsx', '.xls')):
            return self.extract_xlsx_content(url)
        else:
            return None
    
    def extract_xlsx_content(self, xlsx_url):
        """Ekstrak konten dari XLSX dengan penanganan error yang aman."""
        try:
            import openpyxl
            
            # Download ke file sementara
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                response = requests.get(xlsx_url, timeout=15)
                tmp_file.write(response.content)
                tmp_path = tmp_file.name
            
            try:
                # Ekstrak data dari XLSX
                wb = openpyxl.load_workbook(tmp_path, read_only=True)
                text = ""
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    for row in sheet.iter_rows(values_only=True):
                        row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                        text += row_text + "\n"
                return text.strip()
            finally:
                os.unlink(tmp_path)
                
        except ImportError:
            print("⚠️ openpyxl not available. Install it with: pip install openpyxl")
            return None
        except Exception as e:
            print(f"⚠️ XLSX extraction failed for {xlsx_url}: {e}")
            return None